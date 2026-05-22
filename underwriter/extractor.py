"""
Extract underwriter assumptions from a V16.x Restoration Homes spreadsheet.

Returns a dict ready to be passed to `db.upsert_property_from_xlsx()`.
Identifying fields are top-level (address/city/state/zip/county/sqft/listPrice);
the assumption record is the rest of the dict (consumed by `model.compute()`).
"""
from __future__ import annotations

import re
from pathlib import Path
from typing import Any

import openpyxl

from model import DEFAULTS, _default_rental_opex, RENTAL_OPEX_LINES
import title


# ---------------------------------------------------------------------------
# Cell map — Dashboard tab + Standard Comp Comparison
# ---------------------------------------------------------------------------
DASHBOARD_CELLS = {
    "address_line1":           ("Dashboard", "B5"),
    "address_line2":           ("Dashboard", "B6"),

    # ARVs (B = formula/auto, C = override)
    "_bridgeARV_auto":         ("Dashboard", "B23"),
    "_bridgeARV_override":     ("Dashboard", "C23"),
    "_dscrARV_auto":           ("Dashboard", "B24"),
    "_dscrARV_override":       ("Dashboard", "C24"),

    # Rent
    "uwRent":                  ("Dashboard", "B25"),
    "rentOverride":            ("Dashboard", "C25"),

    # Tax & insurance
    "_propTax_auto":           ("Dashboard", "B28"),
    "_propTax_override":       ("Dashboard", "C28"),
    "_insurance_auto":         ("Dashboard", "B29"),
    "_insurance_override":     ("Dashboard", "C29"),

    # Reno
    "_reno_auto":              ("Dashboard", "B22"),
    "_reno_override":          ("Dashboard", "C22"),

    "_purchase_override":      ("Dashboard", "C20"),

    # Bridge / Flip
    "bridgeRate":              ("Dashboard", "B45"),
    "maxLTV":                  ("Dashboard", "B43"),
    "maxLTC":                  ("Dashboard", "B42"),
    "bridgeOrigPct":           ("Dashboard", "B46"),
    "bridgeUWFee":             ("Dashboard", "B47"),
    "bridgeClosingCosts":      ("Dashboard", "B48"),
    "bridgePreClosingExpenses":("Dashboard", "B49"),
    "bridgePostClosingExpenses":("Dashboard","B50"),
    "vacateDays":              ("Dashboard", "F5"),
    "listingRefiDays":         ("Dashboard", "L16"),
    "closingDays":             ("Dashboard", "L17"),
    "utilityCostsMonthly":     ("Dashboard", "O13"),
    "dispoCommissionPct":      ("Dashboard", "O17"),
    "dispoTransferTaxPct":     ("Dashboard", "O18"),
    "dispoClosingCosts":       ("Dashboard", "O20"),
    "usePML":                  ("Dashboard", "F36"),
    "pmlAmount":               ("Dashboard", "F41"),
    "pmlRate":                 ("Dashboard", "F42"),
    "pmlFees":                 ("Dashboard", "F43"),

    # DSCR
    "borrowMaxDSCR":           ("Dashboard", "B33"),
    "dscrRate":                ("Dashboard", "B79"),
    "dscrMaxLTV":              ("Dashboard", "B78"),
    "dscrOriginationPct":      ("Dashboard", "B80"),
    "dscrUWFee":               ("Dashboard", "B81"),
    "dscrClosingCosts":        ("Dashboard", "B82"),
    "prepaidInterestDays":     ("Dashboard", "F76"),
    "hoaMonths":               ("Dashboard", "F77"),
    "insuranceMonths":         ("Dashboard", "F78"),
    "propTaxMonths":           ("Dashboard", "F79"),
    "sellAfterYears":          ("Dashboard", "F83"),
    "amortYears":              ("Dashboard", "F86"),
    "includeAppreciation":     ("Dashboard", "B34"),
    "appreciationRate":        ("Dashboard", "C34"),
    "vacancyRate":             ("Dashboard", "B72"),
    "useDwellaPM":             ("Dashboard", "B36"),

    # Identifying
    "address_full":            ("Standard Comp Comparison", "B3"),
    "county":                  ("Standard Comp Comparison", "B8"),
    "sqft":                    ("Standard Comp Comparison", "B16"),
    "listPrice":               ("Standard Comp Comparison", "E3"),
    "hoa":                     ("Standard Comp Comparison", "B21"),
}

# Row range in the Dashboard's "Rental Operating Expenses (Year 1)" table.
# Each row is (xlsx_label_match, rental_opex_id)
OPEX_ROW_MAP = {
    "HOA":                  "HOA",
    "Solar Panel Pmts":     "solar1",        # first occurrence (annual?)
    "Insurance":            "insurance",
    "Landscaping":          "landscaping",
    "Property Management":  "propMgmt",
    "Repairs & Maintenance":"repairs",
    "Supplies":             "supplies",
    "Property Tax":         "propertyTax",
    "Trash Removal":        "trash",
    "Utilities":            "utilities",
    "Water and Sewer":      "waterSewer",
    "Other Expenses":       "other",
}


def _num(v: Any, default: float = 0.0) -> float:
    if v is None or v == "":
        return default
    try: return float(v)
    except (TypeError, ValueError): return default


def _bool(v: Any, default: bool = False) -> bool:
    if v is None or v == "":
        return default
    if isinstance(v, bool):
        return v
    return str(v).strip().lower() in ("true", "yes", "y", "1")


def _override(auto: Any, override: Any, default: float = 0.0) -> float:
    o = _num(override, None)
    if o not in (None, 0.0):
        return o
    a = _num(auto, None)
    if a not in (None, 0.0):
        return a
    return default


def _extract_opex_table(ws) -> dict:
    """Read rows 54..66 of the Dashboard tab — that's the operating-expense
    section.  Columns:
        A: line label
        C: amount (B=auto, C=override)
        E: 'Annual' / 'Monthly'
        F: annual escalation %
    """
    out = {}
    seen_solar = 0
    for row in range(54, 67):
        label = (ws[f"A{row}"].value or "").strip()
        if not label:
            continue
        amount = _num(ws[f"C{row}"].value)
        freq = (ws[f"E{row}"].value or "Annual").strip()
        esc = _num(ws[f"F{row}"].value)

        # Map label -> rental_opex id
        if label == "Solar Panel Pmts":
            seen_solar += 1
            fid = "solar1" if seen_solar == 1 else "solar2"
        else:
            fid = OPEX_ROW_MAP.get(label)

        if fid:
            out[fid] = {"amount": amount, "frequency": freq, "escalation": esc}
    return out


def extract(xlsx_path: str | Path) -> dict:
    path = Path(xlsx_path)
    wb = openpyxl.load_workbook(path, data_only=True)

    raw: dict[str, Any] = {}
    for key, (sheet_name, cell) in DASHBOARD_CELLS.items():
        if sheet_name not in wb.sheetnames:
            raw[key] = None
            continue
        raw[key] = wb[sheet_name][cell].value

    # ---- identifiers ----
    addr_full = (raw.get("address_full") or "").strip()
    if not addr_full:
        line1 = (raw.get("address_line1") or "").strip()
        line2 = (raw.get("address_line2") or "").strip()
        addr_full = f"{line1}, {line2}".strip(", ")

    street, city, state, zip_code = addr_full, "", "", ""
    m = re.match(r"^(.*?),\s*([^,]+?),?\s*([A-Z]{2})\s*(\d{5})?\s*$",
                 addr_full, re.IGNORECASE)
    if m:
        street = m.group(1).strip()
        city   = m.group(2).strip()
        state  = m.group(3).strip().upper()
        zip_code = (m.group(4) or "").strip()
    else:
        parts = [p.strip() for p in addr_full.split(",")]
        street = parts[0] if parts else addr_full
        if len(parts) >= 3:
            city = parts[1]
            sz = parts[2].split()
            if sz: state = sz[0]
            if len(sz) > 1: zip_code = sz[1]

    # ---- op-ex table ----
    opex = dict(_default_rental_opex(0, 0, 0))
    if "Dashboard" in wb.sheetnames:
        opex.update(_extract_opex_table(wb["Dashboard"]))
    wb.close()

    # ---- assemble ----
    out = dict(DEFAULTS)
    out.update({
        # identifiers
        "address":     addr_full,
        "city":        city,
        "state":       state,
        "zip":         zip_code,
        "county":      (raw.get("county") or "") or "",
        "sqft":        _num(raw.get("sqft"), 0),
        "listPrice":   _num(raw.get("listPrice"), 0),
        "source_file": path.name,

        # valuation
        "bridgeARV":          _override(raw["_bridgeARV_auto"], raw["_bridgeARV_override"]),
        "dscrARV":            _override(raw["_dscrARV_auto"],   raw["_dscrARV_override"]),
        "purchasePrice":      _num(raw["_purchase_override"]),
        "renoBudget":         _override(raw["_reno_auto"], raw["_reno_override"]),
        "uwRent":             _num(raw.get("uwRent")),
        "rentOverride":       _num(raw.get("rentOverride")),
        "annualPropertyTax":  _override(raw["_propTax_auto"], raw["_propTax_override"]),
        "insuranceAnnual":    _override(raw["_insurance_auto"], raw["_insurance_override"]),

        # bridge / dscr (rates / fees / etc.)
        "bridgeRate":         _num(raw.get("bridgeRate"), DEFAULTS["bridgeRate"]),
        "maxLTV":             _num(raw.get("maxLTV"),     DEFAULTS["maxLTV"]),
        "maxLTC":             _num(raw.get("maxLTC"),     DEFAULTS["maxLTC"]),
        "bridgeOrigPct":      _num(raw.get("bridgeOrigPct"), DEFAULTS["bridgeOrigPct"]),
        "bridgeUWFee":        _num(raw.get("bridgeUWFee"),    DEFAULTS["bridgeUWFee"]),
        "bridgeClosingCosts": _num(raw.get("bridgeClosingCosts"), DEFAULTS["bridgeClosingCosts"]),
        "bridgePreClosingExpenses":  _num(raw.get("bridgePreClosingExpenses"),  DEFAULTS["bridgePreClosingExpenses"]),
        "bridgePostClosingExpenses": _num(raw.get("bridgePostClosingExpenses"), DEFAULTS["bridgePostClosingExpenses"]),
        "vacateDays":         _num(raw.get("vacateDays"),       DEFAULTS["vacateDays"]),
        "listingRefiDays":    _num(raw.get("listingRefiDays"),  DEFAULTS["listingRefiDays"]),
        "closingDays":        _num(raw.get("closingDays"),      DEFAULTS["closingDays"]),
        "utilityCostsMonthly":_num(raw.get("utilityCostsMonthly"), DEFAULTS["utilityCostsMonthly"]),
        "dispoCommissionPct": _num(raw.get("dispoCommissionPct"),  DEFAULTS["dispoCommissionPct"]),
        "dispoTransferTaxPct":_num(raw.get("dispoTransferTaxPct"), DEFAULTS["dispoTransferTaxPct"]),
        "dispoClosingCosts":  _num(raw.get("dispoClosingCosts"),   DEFAULTS["dispoClosingCosts"]),
        "usePML":             _bool(raw.get("usePML"),    DEFAULTS["usePML"]),
        "pmlAmount":          _num(raw.get("pmlAmount"),  DEFAULTS["pmlAmount"]),
        "pmlRate":            _num(raw.get("pmlRate"),    DEFAULTS["pmlRate"]),
        "pmlFees":            _num(raw.get("pmlFees"),    DEFAULTS["pmlFees"]),
        "borrowMaxDSCR":      _bool(raw.get("borrowMaxDSCR"), DEFAULTS["borrowMaxDSCR"]),
        "dscrRate":           _num(raw.get("dscrRate"),       DEFAULTS["dscrRate"]),
        "dscrMaxLTV":         _num(raw.get("dscrMaxLTV"),     DEFAULTS["dscrMaxLTV"]),
        "dscrOriginationPct": _num(raw.get("dscrOriginationPct"), DEFAULTS["dscrOriginationPct"]),
        "dscrUWFee":          _num(raw.get("dscrUWFee"),      DEFAULTS["dscrUWFee"]),
        "dscrClosingCosts":   _num(raw.get("dscrClosingCosts"), DEFAULTS["dscrClosingCosts"]),
        "prepaidInterestDays":_num(raw.get("prepaidInterestDays"), DEFAULTS["prepaidInterestDays"]),
        "hoaMonths":          _num(raw.get("hoaMonths"),      DEFAULTS["hoaMonths"]),
        "insuranceMonths":    _num(raw.get("insuranceMonths"),DEFAULTS["insuranceMonths"]),
        "propTaxMonths":      _num(raw.get("propTaxMonths"),  DEFAULTS["propTaxMonths"]),
        "sellAfterYears":     _num(raw.get("sellAfterYears"), DEFAULTS["sellAfterYears"]),
        "amortYears":         _num(raw.get("amortYears"),     DEFAULTS["amortYears"]),
        "includeAppreciation":_bool(raw.get("includeAppreciation"), DEFAULTS["includeAppreciation"]),
        "appreciationRate":   _num(raw.get("appreciationRate"),DEFAULTS["appreciationRate"]),
        "vacancyRate":        _num(raw.get("vacancyRate"),    DEFAULTS["vacancyRate"]),
        "useDwellaPM":        _bool(raw.get("useDwellaPM"),   DEFAULTS["useDwellaPM"]),
    })

    # ---- sync the auto-from op-ex lines (insurance, property tax) ----
    opex.setdefault("insurance",    {})["amount"]   = out["insuranceAnnual"]
    opex.setdefault("insurance",    {}).setdefault("frequency", "Annual")
    opex.setdefault("insurance",    {}).setdefault("escalation", 0.03)
    opex.setdefault("propertyTax",  {})["amount"]   = out["annualPropertyTax"]
    opex.setdefault("propertyTax",  {}).setdefault("frequency", "Annual")
    opex.setdefault("propertyTax",  {}).setdefault("escalation", 0.03)

    # PM defaults to 10% of rent if not explicitly set in xlsx
    if not opex.get("propMgmt", {}).get("amount"):
        opex["propMgmt"] = {"amount": (out["rentOverride"] or out["uwRent"]) * 0.10,
                             "frequency": "Monthly", "escalation": 0.03}

    out["rentalOpEx"] = opex

    # ---- prelim title (best-effort) ----------------------------------------
    pt = _extract_prelim_title(openpyxl.load_workbook(path, data_only=True))
    out["prelim_title"] = pt
    return out


def _date_str(v) -> str:
    """Convert datetime/date/string to ISO yyyy-mm-dd or empty."""
    if v in (None, ""):
        return ""
    if hasattr(v, "isoformat"):
        try:
            return v.date().isoformat() if hasattr(v, "date") else v.isoformat()
        except Exception:
            return str(v)
    return str(v).split(" ")[0]


def _extract_prelim_title(wb) -> dict:
    """Pull intake-form data from the workbook's Prelim. Title sheet."""
    if "Prelim. Title" not in wb.sheetnames:
        return dict(title.DEFAULT_PRELIM_TITLE)

    ws = wb["Prelim. Title"]
    out = {**title.DEFAULT_PRELIM_TITLE}

    # Section 1: Property Details
    out["parcelId"] = str(ws["B4"].value or "").strip()
    out["owners"]   = str(ws["B6"].value or "").strip()

    # Section 2: 1st mortgage from row 15 (or 24 = current mortgage)
    out["mortgage1"] = {
        "company":            str(ws["D15"].value or "").strip(),
        "date":               _date_str(ws["C15"].value),
        "initialAmount":      _num(ws["H15"].value),
        "rate":               _num(ws["I15"].value),
        "assignmentServicer": str(ws["D24"].value or "").strip(),
        "assignmentDate":     _date_str(ws["C24"].value),
    }

    # Section 3: Payoff statement / foreclosure order (rows 45-58)
    payoff1 = title.empty_payoff()
    payoff1["statementDate"]      = _date_str(ws["F45"].value)
    payoff1["currentPrincipal"]   = _num(ws["H45"].value)
    payoff1["cumulativeInterest"] = _num(ws["H46"].value)
    payoff1["taxesOwed"]          = _num(ws["H50"].value)
    payoff1["insuranceOwed"]      = _num(ws["H51"].value)
    payoff1["escrowsOwed"]        = _num(ws["H52"].value)
    payoff1["lateFees"]           = _num(ws["H53"].value)
    payoff1["foreclosureCosts"]   = _num(ws["H55"].value)
    payoff1["attorneyFees"]       = _num(ws["H56"].value)
    payoff1["other"]              = _num(ws["H57"].value)
    out["payoff1"] = payoff1

    # Section 5: Liens (rows 27-29 in spreadsheet's 2) Property Lien Index)
    liens = []
    for r in (27, 28, 29):
        amt = _num(ws[f"H{r}"].value)
        if amt or ws[f"D{r}"].value or ws[f"C{r}"].value:
            liens.append({
                "dateFiled": _date_str(ws[f"C{r}"].value),
                "holder":    str(ws[f"D{r}"].value or "").strip(),
                "lienNumber":str(ws[f"E{r}"].value or "").strip(),
                "bookPage":  str(ws[f"F{r}"].value or "").strip(),
                "principalAmount": amt,
                "interestRate":    _num(ws[f"I{r}"].value, 0.0625) or 0.0625,
            })
    while len(liens) < 3:
        liens.append(title.empty_lien())
    out["liens"] = liens

    # Section 6: Judgments (rows 35-39)
    judgments = []
    for r in (35, 36, 37, 38, 39):
        amt = _num(ws[f"H{r}"].value)
        if amt or (ws[f"D{r}"].value and "Creditor" not in str(ws[f"D{r}"].value)):
            judgments.append({
                "judgmentDate": _date_str(ws[f"C{r}"].value),
                "plaintiff":    str(ws[f"D{r}"].value or "").strip(),
                "caseNumber":   str(ws[f"E{r}"].value or "").strip(),
                "principalAmount": amt,
                "interestRate":    _num(ws[f"I{r}"].value, 0.08) or 0.08,
            })
    while len(judgments) < 3:
        judgments.append(title.empty_judgment())
    out["judgments"] = judgments

    return out


if __name__ == "__main__":
    import json
    import sys
    p = sys.argv[1] if len(sys.argv) > 1 else "properties/RH Underwriter V16.2 [379 Curtis Dr, Sumter, SC 29153] Updated 9-25-2025.xlsx"
    print(json.dumps(extract(p), indent=2, default=str))
